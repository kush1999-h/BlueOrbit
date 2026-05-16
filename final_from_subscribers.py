"""Produce merged CSV starting from subscribers as the base.
For each SID,SUBSCRIBERSN in subscriber file:
 - include SUBSCRIBERIDENTIFIER,HOMESRVZONE
 - for each subscription (if any): map service->quota(s) and attach quota details
 - if no subscriptions, write one row with blanks for service/quota
"""
import csv
import os
import shutil
import sys

BASE_DIR = os.path.dirname(__file__)
SUBSCRIBER_F = os.path.join(BASE_DIR, 'UPCC_1840_20260510130411_1_SUBSCRIBER.txt')
SUBSCRIPTION_F = os.path.join(BASE_DIR, 'UPCC_1840_20260510130411_1_SUBSCRIPTION.txt')
SUBQUOTA_F = os.path.join(BASE_DIR, 'UPCC_1840_20260510130411_1_SUBQUOTA.txt')
XLSX_F = os.path.join(BASE_DIR, 'UPCC03_service_quota.xlsx')
OUT_F = os.path.join(BASE_DIR, 'merged_SERVICE_TO_QUOTA_from_subscribers.csv')
OUT_TXT_F = os.path.join(BASE_DIR, 'merged_SERVICE_TO_QUOTA_from_subscribers.txt')

try:
    import openpyxl
except Exception as e:
    print('Missing openpyxl:', e)
    raise

# load service->quota mapping
wb = openpyxl.load_workbook(XLSX_F, data_only=True)
ws = wb.active
header = [str(cell.value).strip() if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
svc_col = None
quota_col = None
for idx, h in enumerate(header):
    hlow = h.lower()
    if 'service' in hlow and svc_col is None:
        svc_col = idx
    if 'quota' in hlow and quota_col is None:
        quota_col = idx
if svc_col is None or quota_col is None:
    print('Could not detect SERVICE/QUOTA columns in', XLSX_F)
    print('Headers:', header)
    sys.exit(2)
service_to_quota = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    svc = row[svc_col]
    q = row[quota_col]
    if svc is None:
        continue
    svc = str(svc).strip()
    if not svc:
        continue
    if q is None:
        continue
    parts = [p.strip() for p in str(q).replace(';', ',').split(',') if p.strip()]
    service_to_quota.setdefault(svc, []).extend(parts)
print('Loaded service->quota mappings:', len(service_to_quota))

# load subscribers
subscriber_map = {}
with open(SUBSCRIBER_F, newline='', encoding='utf-8', errors='replace') as fh:
    reader = csv.DictReader(fh)
    for row in reader:
        key = (row.get('SID','').strip(), row.get('SUBSCRIBERSN','').strip())
        subscriber_map[key] = {
            'SUBSCRIBERIDENTIFIER': row.get('SUBSCRIBERIDENTIFIER',''),
            'HOMESRVZONE': row.get('HOMESRVZONE','')
        }
print('Loaded subscribers:', len(subscriber_map))

# load subscriptions grouped by subscriber key
subs_by_sub = {}
with open(SUBSCRIPTION_F, newline='', encoding='utf-8', errors='replace') as fh:
    reader = csv.DictReader(fh)
    for row in reader:
        key = (row.get('SID','').strip(), row.get('SUBSCRIBERSN','').strip())
        subs_by_sub.setdefault(key, []).append({
            'SERVICENAME': row.get('SERVICENAME',''),
            'SUBSCRIBEDATETIME': row.get('SUBSCRIBEDATETIME',''),
            'VALIDFROMDATETIME': row.get('VALIDFROMDATETIME',''),
            'EXPIREDATETIME': row.get('EXPIREDATETIME','')
        })
print('Loaded subscription groups:', len(subs_by_sub))

# load quotas keyed by (SID,SUBSCRIBERSN,QUOTANAME)
quota_map = {}
with open(SUBQUOTA_F, newline='', encoding='utf-8', errors='replace') as fh:
    reader = csv.DictReader(fh)
    for row in reader:
        key = (row.get('SID','').strip(), row.get('SUBSCRIBERSN','').strip(), row.get('QUOTANAME','').strip())
        quota_map.setdefault(key, []).append(row)
print('Loaded quota rows:', sum(len(v) for v in quota_map.values()))

# output
headers = ['SID','SUBSCRIBERSN','SUBSCRIBERIDENTIFIER','HOMESRVZONE','SERVICENAME','SUBSCRIBEDATETIME','VALIDFROMDATETIME','EXPIREDATETIME','QUOTANAME','NEXTRESETDATETIME','CONSUMPTION']
rows_out = 0
with open(OUT_F, 'w', newline='', encoding='utf-8') as outf:
    writer = csv.DictWriter(outf, fieldnames=headers)
    writer.writeheader()
    for key, subinfo in subscriber_map.items():
        sid, subsn = key
        subscriptions = subs_by_sub.get(key, [])
        if not subscriptions:
            # write one row with subscriber info and blanks
            out = {'SID': sid, 'SUBSCRIBERSN': subsn, 'SUBSCRIBERIDENTIFIER': subinfo.get('SUBSCRIBERIDENTIFIER',''), 'HOMESRVZONE': subinfo.get('HOMESRVZONE',''),
                   'SERVICENAME': '', 'SUBSCRIBEDATETIME':'', 'VALIDFROMDATETIME':'', 'EXPIREDATETIME':'', 'QUOTANAME':'', 'NEXTRESETDATETIME':'', 'CONSUMPTION':''}
            writer.writerow(out)
            rows_out += 1
            continue
        for s in subscriptions:
            svc = s.get('SERVICENAME','')
            mapped = service_to_quota.get(svc, [])
            if not mapped:
                out = {'SID': sid, 'SUBSCRIBERSN': subsn, 'SUBSCRIBERIDENTIFIER': subinfo.get('SUBSCRIBERIDENTIFIER',''), 'HOMESRVZONE': subinfo.get('HOMESRVZONE',''),
                       'SERVICENAME': svc, 'SUBSCRIBEDATETIME': s.get('SUBSCRIBEDATETIME',''), 'VALIDFROMDATETIME': s.get('VALIDFROMDATETIME',''), 'EXPIREDATETIME': s.get('EXPIREDATETIME',''),
                       'QUOTANAME':'', 'NEXTRESETDATETIME':'', 'CONSUMPTION':''}
                writer.writerow(out)
                rows_out += 1
                continue
            for qname in mapped:
                qkey = (sid, subsn, qname)
                qrows = quota_map.get(qkey)
                if qrows:
                    for qr in qrows:
                        out = {'SID': sid, 'SUBSCRIBERSN': subsn, 'SUBSCRIBERIDENTIFIER': subinfo.get('SUBSCRIBERIDENTIFIER',''), 'HOMESRVZONE': subinfo.get('HOMESRVZONE',''),
                               'SERVICENAME': svc, 'SUBSCRIBEDATETIME': s.get('SUBSCRIBEDATETIME',''), 'VALIDFROMDATETIME': s.get('VALIDFROMDATETIME',''), 'EXPIREDATETIME': s.get('EXPIREDATETIME',''),
                               'QUOTANAME': qname, 'NEXTRESETDATETIME': qr.get('NEXTRESETDATETIME',''), 'CONSUMPTION': qr.get('CONSUMPTION','')}
                        writer.writerow(out)
                        rows_out += 1
                else:
                    out = {'SID': sid, 'SUBSCRIBERSN': subsn, 'SUBSCRIBERIDENTIFIER': subinfo.get('SUBSCRIBERIDENTIFIER',''), 'HOMESRVZONE': subinfo.get('HOMESRVZONE',''),
                           'SERVICENAME': svc, 'SUBSCRIBEDATETIME': s.get('SUBSCRIBEDATETIME',''), 'VALIDFROMDATETIME': s.get('VALIDFROMDATETIME',''), 'EXPIREDATETIME': s.get('EXPIREDATETIME',''),
                           'QUOTANAME': qname, 'NEXTRESETDATETIME': '', 'CONSUMPTION': ''}
                    writer.writerow(out)
                    rows_out += 1

print('Wrote', OUT_F, 'rows:', rows_out)
shutil.copyfile(OUT_F, OUT_TXT_F)
print('Wrote', OUT_TXT_F, 'rows:', rows_out)
